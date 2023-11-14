import tkinter as tk
from tkinter import filedialog, ttk
import pandas as pd
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np


def select_file_and_sheet():
    def on_file_select():
        nonlocal file_path
        file_path = filedialog.askopenfilename(
            title="Zieldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            df = pd.ExcelFile(file_path)
            sheet_combo["values"] = df.sheet_names
            sheet_combo.current(0)
            select_button["state"] = "normal"

    def on_select():
        nonlocal selected_sheet
        selected_sheet = sheet_combo.get()

        root.destroy()

    root = tk.Tk()

    root.title("Excel-Datei und Arbeitsblatt auswählen")

    file_path = None
    selected_sheet = None

    tk.Label(root, text="Wählen Sie die Excel-Datei:").pack(padx=10, pady=5)
    file_button = tk.Button(root, text="Datei auswählen", command=on_file_select)
    file_button.pack(padx=10, pady=5)

    tk.Label(root, text="Wählen Sie das Arbeitsblatt:").pack(padx=10, pady=5)
    sheet_combo = ttk.Combobox(root, state="readonly")
    sheet_combo.pack(padx=10, pady=5)

    select_button = tk.Button(
        root, text="Auswählen", state="disabled", command=on_select
    )
    select_button.pack(padx=10, pady=10)

    root.mainloop()

    return file_path, selected_sheet


def select_unique_identifier(file_path, sheet_name):
    def on_select():
        nonlocal selected_identifier
        selected_identifier = identifier_combo.get()
        root.destroy()

    root = tk.Tk()
    root.title("Einzigartigen Identifikator auswählen")

    df = pd.read_excel(file_path, sheet_name=sheet_name)
    column_names = df.columns.tolist()

    tk.Label(root, text="Wählen Sie den einzigartigen Identifikator:").pack(
        padx=10, pady=5
    )
    identifier_combo = ttk.Combobox(root, values=column_names, state="readonly")
    identifier_combo.pack(padx=10, pady=5)
    identifier_combo.current(0)

    selected_identifier = None
    select_button = tk.Button(root, text="Auswählen", command=on_select)
    select_button.pack(padx=10, pady=10)

    root.mainloop()

    return selected_identifier


def select_source_file_and_sheet(target_columns):
    def on_file_select():
        nonlocal source_file_path
        source_file_path = filedialog.askopenfilename(
            title="Quelldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if source_file_path:
            df = pd.ExcelFile(source_file_path)
            sheet_combo["values"] = df.sheet_names
            sheet_combo.current(0)
            select_button["state"] = "normal"

    def on_select():
        nonlocal source_selected_sheet
        source_selected_sheet = sheet_combo.get()

        root.destroy()

    root = tk.Tk()
    root.title("Quelldatei und Arbeitsblatt auswählen")

    source_file_path = None
    source_selected_sheet = None

    tk.Label(root, text="Wählen Sie die Quelldatei:").pack(padx=10, pady=5)
    file_button = tk.Button(root, text="Datei auswählen", command=on_file_select)
    file_button.pack(padx=10, pady=5)

    tk.Label(root, text="Wählen Sie das Arbeitsblatt:").pack(padx=10, pady=5)
    sheet_combo = ttk.Combobox(root, state="readonly")
    sheet_combo.pack(padx=10, pady=5)

    select_button = tk.Button(
        root, text="Auswählen", state="disabled", command=on_select
    )
    select_button.pack(padx=10, pady=10)

    root.mainloop()

    # Überprüfen der Spaltenübereinstimmung
    source_df = pd.read_excel(source_file_path, sheet_name=source_selected_sheet)
    source_columns = source_df.columns.tolist()
    missing_columns = [col for col in target_columns if col not in source_columns]
    extra_columns = [col for col in source_columns if col not in target_columns]
    if missing_columns or extra_columns:
        message = "In der Quelldatei fehlen folgende Spalten: {}\nZusätzliche Spalten in der Quelldatei: {}\nMöchten Sie trotzdem fortfahren?".format(
            missing_columns, extra_columns
        )
        continue_comparison = messagebox.askyesno("Spaltenüberprüfung", message)
        if not continue_comparison:
            return None, None

    return source_file_path, source_selected_sheet


def is_equal(val1, val2):
    if pd.isna(val1) and pd.isna(val2):
        return True
    if pd.isna(val1) and val2 == "":
        return True
    if val1 == "" and pd.isna(val2):
        return True
    return val1 == val2


def compare_and_update(
    target_file, target_sheet, source_file, source_sheet, unique_identifier
):
    # Excel-Dateien einlesen
    target_df = pd.read_excel(target_file, sheet_name=target_sheet)
    source_df = pd.read_excel(source_file, sheet_name=source_sheet)

    # Mapping für schnellen Zugriff erstellen
    source_map = source_df.set_index(unique_identifier).to_dict(orient="index")

    # openpyxl Workbook und Worksheet für die Aktualisierung vorbereiten
    book = openpyxl.load_workbook(target_file)
    sheet = book[target_sheet]
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in sheet[1]:
        cell.fill = blue_fill

    # Überschrift für die "Check"-Spalte hinzufügen
    sheet.cell(row=1, column=len(target_df.columns) + 1, value="Check").fill = red_fill

    for i, row in target_df.iterrows():
        row_id = row[unique_identifier]
        row_changed = False
        if row_id in source_map:
            for j, col in enumerate(target_df.columns):
                if not is_equal(row[col], source_map[row_id].get(col, row[col])):
                    # Aktualisiere die Zelle und färbe sie orange
                    cell = sheet.cell(row=i + 2, column=j + 1)
                    cell.value = source_map[row_id][col]
                    cell.fill = orange_fill
                    row_changed = True
            if row_changed:
                # Markiere die Zeile als "geändert"
                sheet.cell(
                    row=i + 2, column=len(target_df.columns) + 1, value="changed"
                )
            else:
                # Markiere die Zeile als "original"
                sheet.cell(
                    row=i + 2, column=len(target_df.columns) + 1, value="original"
                )

        else:
            # Färbe die Zeile rot für "gelöscht"
            for col in range(1, len(target_df.columns) + 2):
                sheet.cell(row=i + 2, column=col).fill = red_fill
            sheet.cell(row=i + 2, column=len(target_df.columns) + 1, value="deleted")

    # Neue Zeilen aus der Quelle hinzufügen und grün einfärben
    for row_id in set(source_df[unique_identifier]) - set(target_df[unique_identifier]):
        new_row = source_df[source_df[unique_identifier] == row_id].iloc[0]
        sheet.append(new_row.tolist() + ["new"])
        for col in range(1, len(new_row) + 2):
            sheet.cell(row=sheet.max_row, column=col).fill = green_fill

    # Spaltenbreite anpassen
    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        adjusted_width = max_length + 2
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column_cells[0].column)
        ].width = adjusted_width

    # Speichere die aktualisierte Datei
    updated_file = target_file.replace(".xlsx", "_updated.xlsx")
    book.save(updated_file)

    return updated_file


def main():
    # Erster Schritt: Ziel-Datei und Arbeitsblatt auswählen
    target_file, target_sheet = select_file_and_sheet()
    if not target_file or not target_sheet:
        # print("Zieldatei oder Arbeitsblatt wurde nicht ausgewählt.")
        return

    # print("Ausgewählte Datei:", target_file)
    # print("Ausgewähltes Arbeitsblatt:", target_sheet)

    # Zweiter Schritt: Einzigartigen Identifikator auswählen
    print("Einzigartigen Identifikator auswählen")
    unique_identifier = select_unique_identifier(target_file, target_sheet)
    if not unique_identifier:
        # print("Einzigartiger Identifikator wurde nicht ausgewählt.")
        return

    print("Ausgewählter einzigartiger Identifikator:", unique_identifier)

    # Dritter Schritt: Quelldatei und Arbeitsblatt auswählen
    target_df = pd.read_excel(target_file, sheet_name=target_sheet)
    target_sheet_columns = target_df.columns.tolist()

    source_file, source_sheet = select_source_file_and_sheet(target_sheet_columns)
    if not source_file or not source_sheet:
        # print("Quelldatei oder Arbeitsblatt wurde nicht ausgewählt.")
        return

    # print("Ausgewählte Quelldatei:", source_file)
    # print("Ausgewähltes Arbeitsblatt in der Quelldatei:", source_sheet)

    # Vierter Schritt: Daten vergleichen und aktualisieren
    updated_file = compare_and_update(
        target_file, target_sheet, source_file, source_sheet, unique_identifier
    )

    print("Aktualisierte Datei gespeichert unter:", updated_file)
    if __name__ != "__main__":
        import main

        return updated_file


# Hauptfunktion aufrufen
if __name__ == "__main__":
    main()
