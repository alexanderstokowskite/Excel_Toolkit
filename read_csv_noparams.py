import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill


def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width


def save_as_excel(csv_file_path):
    try:
        df = pd.read_csv(csv_file_path)
        excel_file_path = csv_file_path.replace(".csv", ".xlsx")

        # Save DataFrame to Excel
        df.to_excel(excel_file_path, index=False, engine="openpyxl")

        # Open the Excel file to adjust formatting
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active

        adjust_column_width(sheet)

        # Apply header formatting if there is a header
        if df.columns is not None:
            header_fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )
            for cell in sheet[1]:
                cell.fill = header_fill

        workbook.save(excel_file_path)
        messagebox.showinfo(
            "Erfolg", f"Datei erfolgreich als '{excel_file_path}' gespeichert."
        )

    except FileNotFoundError:
        messagebox.showerror(
            "Fehler", f"Die Datei '{csv_file_path}' wurde nicht gefunden."
        )
    except pd.errors.EmptyDataError:
        messagebox.showerror("Fehler", f"Die Datei '{csv_file_path}' ist leer.")
    except pd.errors.ParserError as e:
        messagebox.showerror("Fehler", f"Fehler beim Parsen der CSV-Datei: {e}")
    except Exception as e:
        messagebox.showerror(
            "Unerwarteter Fehler", f"Ein unerwarteter Fehler ist aufgetreten: {e}"
        )


def select_and_convert_file():
    root = tk.Tk()
    root.withdraw()  # Verstecke das Hauptfenster

    file_path = filedialog.askopenfilename(filetypes=[("CSV-Dateien", "*.csv")])
    if file_path:
        save_as_excel(file_path)


if __name__ == "__main__":
    select_and_convert_file()
