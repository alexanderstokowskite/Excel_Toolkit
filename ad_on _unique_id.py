import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def copy_and_update(
    target_file, target_sheet, source_file, source_sheet, unique_identifier
):
    # Excel-Dateien einlesen
    target_df = pd.read_excel(target_file, sheet_name=target_sheet)
    source_df = pd.read_excel(source_file, sheet_name=source_sheet)

    # Mapping für schnellen Zugriff erstellen
    source_map = source_df.set_index(unique_identifier).to_dict(orient="index")

    # openpyxl Workbook und Worksheet für die Aktualisierung vorbereiten
    book = openpyxl.load_workbook(target_file)
    sheet = book[target_sheet]
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )

    # Ziel- und Quellspalten synchronisieren
    source_columns = [col for col in source_df.columns if col != unique_identifier]

    # Neue Spalten in die Zieldatei einfügen
    for idx, col in enumerate(source_columns):
        col_index = len(target_df.columns) + idx + 1
        sheet.cell(row=1, column=col_index, value=col)

    for i, row in target_df.iterrows():
        row_id = row[unique_identifier]
        if row_id in source_map:
            for idx, col in enumerate(source_columns):
                if col in source_map[row_id]:
                    col_index = len(target_df.columns) + idx + 1  # 1-basiert
                    cell = sheet.cell(row=i + 2, column=col_index)
                    cell.value = source_map[row_id][col]
                    cell.fill = orange_fill

    # Spaltenbreite anpassen
    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        adjusted_width = max_length + 2
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column_cells[0].column)
        ].width = adjusted_width

    # Speichere die aktualisierte Datei
    updated_file = target_file.replace(".xlsx", "_updated.xlsx")
    book.save(updated_file)

    return updated_file


def gather_inputs():
    def on_target_file_select():
        file_path = filedialog.askopenfilename(
            title="Zieldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            df = pd.ExcelFile(file_path)
            target_sheet_combo["values"] = df.sheet_names
            target_sheet_combo.current(0)
            target_file_entry.delete(0, tk.END)
            target_file_entry.insert(0, file_path)

            # Lese die Ziel-Datei ein und extrahiere die Spaltennamen
            target_df = pd.read_excel(file_path, sheet_name=target_sheet_combo.get())
            unique_id_combo["values"] = target_df.columns.tolist()
            unique_id_combo.current(0)

    def on_source_file_select():
        file_path = filedialog.askopenfilename(
            title="Quelldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            df = pd.ExcelFile(file_path)
            source_sheet_combo["values"] = df.sheet_names
            source_sheet_combo.current(0)
            source_file_entry.delete(0, tk.END)
            source_file_entry.insert(0, file_path)

    def on_confirm():
        # Sammelt die Daten und schließt den Dialog
        user_inputs = {
            "target_file": target_file_entry.get(),
            "target_sheet": target_sheet_combo.get(),
            "source_file": source_file_entry.get(),
            "source_sheet": source_sheet_combo.get(),
            "unique_identifier": unique_id_combo.get(),
        }
        dialog.user_inputs = user_inputs  # Speichere die Eingaben im Dialog-Objekt
        dialog.destroy()  # Schließt den Dialog

    dialog = tk.Toplevel()
    dialog.title("File & Sheet Selector")
    dialog.geometry("420x220+1000+300")
    dialog.configure(bg="grey")  # Setzt die Hintergrundfarbe des Dialogs

    dialog.user_inputs = None  # Wird später mit den Benutzereingaben gefüllt

    # Zieldatei und Arbeitsblatt
    tk.Label(dialog, text="Target file:", bg="light green", width=15).grid(
        row=0, column=0, sticky="w", padx=10, pady=5
    )
    target_file_entry = tk.Entry(dialog, width=23)
    target_file_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(dialog, text="Select Target", command=on_target_file_select).grid(
        row=0, column=2, padx=10, pady=5
    )

    # Ziel-Arbeitsblatt
    tk.Label(dialog, text="Target Sheet:", bg="light green", width=15).grid(
        row=1, column=0, sticky="w", padx=10, pady=5
    )
    target_sheet_combo = ttk.Combobox(dialog, state="readonly", width=20)
    target_sheet_combo.grid(row=1, column=1, padx=10, pady=5)

    # Quelldatei und Arbeitsblatt
    tk.Label(dialog, text="Source File:", bg="orange", width=15).grid(
        row=2, column=0, sticky="w", padx=10, pady=5
    )
    source_file_entry = tk.Entry(dialog, width=23)
    source_file_entry.grid(row=2, column=1, padx=10, pady=5)
    tk.Button(dialog, text="Select Source", command=on_source_file_select).grid(
        row=2, column=2, padx=10, pady=5
    )

    # Quell-Arbeitsblatt
    tk.Label(dialog, text="Source Sheet:", bg="orange", width=15).grid(
        row=3, column=0, sticky="w", padx=10, pady=5
    )
    source_sheet_combo = ttk.Combobox(dialog, state="readonly", width=20)
    source_sheet_combo.grid(row=3, column=1, padx=10, pady=5)

    # Einzigartiger Identifikator
    tk.Label(dialog, text="Unique Identifier:", bg="light blue", width=15).grid(
        row=4, column=0, sticky="w", padx=10, pady=5
    )
    unique_id_combo = ttk.Combobox(dialog, state="readonly")
    unique_id_combo.grid(row=4, column=1, padx=10, pady=5)

    # Bestätigungsknopf
    tk.Button(dialog, text="Confirm & Execute", command=on_confirm).grid(
        row=5, column=1, padx=10, pady=10
    )

    dialog.wait_window()  # Wartet, bis der Dialog geschlossen wird
    return dialog.user_inputs  # Gibt die Benutzereingaben zurück


def main():
    inputs = gather_inputs()

    if inputs is None:
        print("Prozess abgebrochen.")
        return

    updated_file = copy_and_update(
        inputs["target_file"],
        inputs["target_sheet"],
        inputs["source_file"],
        inputs["source_sheet"],
        inputs["unique_identifier"],
    )

    print("Aktualisierte Datei gespeichert unter:", updated_file)


# Hauptfunktion aufrufen
if __name__ == "__main__":
    main()
