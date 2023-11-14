import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np


def is_equal(val1, val2):
    if pd.isna(val1) and pd.isna(val2):
        return True
    if pd.isna(val1) and val2 == "":
        return True
    if val1 == "" and pd.isna(val2):
        return True
    return val1 == val2


def gather_inputs():
    def on_target_file_select():
        file_path = filedialog.askopenfilename(
            title="Zieldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            df = pd.ExcelFile(file_path)
            target_sheet_combo["values"] = df.sheet_names
            target_sheet_combo.current(0)
            target_file_entry.delete(0, tk.END)
            target_file_entry.insert(0, file_path)

    def on_source_file_select():
        file_path = filedialog.askopenfilename(
            title="Quelldatei auswählen", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            df = pd.ExcelFile(file_path)
            source_sheet_combo["values"] = df.sheet_names
            source_sheet_combo.current(0)
            source_file_entry.delete(0, tk.END)
            source_file_entry.insert(0, file_path)

    def on_confirm():
        # Sammelt die Daten und schließt den Dialog
        user_inputs = {
            "target_file": target_file_entry.get(),
            "target_sheet": target_sheet_combo.get(),
            "source_file": source_file_entry.get(),
            "source_sheet": source_sheet_combo.get(),
        }
        dialog.user_inputs = user_inputs  # Speichere die Eingaben im Dialog-Objekt
        dialog.destroy()  # Schließt den Dialog

    dialog = tk.Toplevel()
    dialog.title("File & Sheet Selector")
    dialog.geometry("420x220+1000+300")
    dialog.configure(bg="grey")  # Setzt die Hintergrundfarbe des Dialogs

    dialog.user_inputs = None  # Wird später mit den Benutzereingaben gefüllt

    # Zieldatei und Arbeitsblatt
    tk.Label(dialog, text="Target file:", bg="light green", width=15).grid(
        row=0, column=0, sticky="w", padx=10, pady=5
    )
    target_file_entry = tk.Entry(dialog, width=23)
    target_file_entry.grid(row=0, column=1, padx=10, pady=5)
    tk.Button(dialog, text="Select Target", command=on_target_file_select).grid(
        row=0, column=2, padx=10, pady=5
    )

    # Ziel-Arbeitsblatt
    tk.Label(dialog, text="Target Sheet:", bg="light green", width=15).grid(
        row=1, column=0, sticky="w", padx=10, pady=5
    )
    target_sheet_combo = ttk.Combobox(dialog, state="readonly", width=20)
    target_sheet_combo.grid(row=1, column=1, padx=10, pady=5)

    # Quelldatei und Arbeitsblatt
    tk.Label(dialog, text="Source File:", bg="orange", width=15).grid(
        row=2, column=0, sticky="w", padx=10, pady=5
    )
    source_file_entry = tk.Entry(dialog, width=23)
    source_file_entry.grid(row=2, column=1, padx=10, pady=5)
    tk.Button(dialog, text="Select Source", command=on_source_file_select).grid(
        row=2, column=2, padx=10, pady=5
    )

    # Quell-Arbeitsblatt
    tk.Label(dialog, text="Source Sheet:", bg="orange", width=15).grid(
        row=3, column=0, sticky="w", padx=10, pady=5
    )
    source_sheet_combo = ttk.Combobox(dialog, state="readonly", width=20)
    source_sheet_combo.grid(row=3, column=1, padx=10, pady=5)

    # Bestätigungsknopf
    tk.Button(dialog, text="Confirm & Execute", command=on_confirm).grid(
        row=5, column=1, padx=10, pady=10
    )

    dialog.wait_window()  # Wartet, bis der Dialog geschlossen wird
    return dialog.user_inputs  # Gibt die Benutzereingaben zurück


def compare_rows_and_update(target_file, target_sheet, source_file, source_sheet):
    # Excel-Dateien einlesen
    target_df = pd.read_excel(target_file, sheet_name=target_sheet)
    source_df = pd.read_excel(source_file, sheet_name=source_sheet)

    # Anzahl der Zeilen überprüfen
    if len(target_df) != len(source_df):
        message = "Die Anzahl der Zeilen in den Dateien stimmt nicht überein. Möchten Sie trotzdem fortfahren?"
        if not messagebox.askyesno("Zeilenanzahl-Überprüfung", message):
            return None

    # openpyxl Workbook und Worksheet für die Aktualisierung vorbereiten
    book = openpyxl.load_workbook(target_file)
    sheet = book[target_sheet]
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in sheet[1]:
        cell.fill = blue_fill

    # Überschrift für die "Check"-Spalte hinzufügen
    sheet.cell(row=1, column=len(target_df.columns) + 1, value="Check").fill = red_fill
    try:
        # Zeilenweise Vergleich
        for i in range(len(target_df)):
            row_changed = False
            for col in target_df.columns:
                if not is_equal(target_df.at[i, col], source_df.at[i, col]):
                    sheet.cell(
                        row=i + 2,
                        column=target_df.columns.get_loc(col) + 1,
                        value=source_df.at[i, col],
                    )
                    sheet.cell(
                        row=i + 2, column=target_df.columns.get_loc(col) + 1
                    ).fill = orange_fill
                    row_changed = True
            sheet.cell(
                row=i + 2,
                column=len(target_df.columns) + 1,
                value="changed" if row_changed else "original",
            )
    except Exception as e:
        # messagebox.showerror("Fehler", str(e))
        print(
            f"Fehler bei der Verarbeitung der Zeile {i}. Weiter mit der nächsten Zeile."
        )

    # Spaltenbreite anpassen
    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        adjusted_width = max_length + 2
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column_cells[0].column)
        ].width = adjusted_width

    status_column_index = (
        len(target_df.columns) + 1
    )  # Die Spalte, in der "Status" steht, ist eine Spalte nach den Daten
    sheet.auto_filter.ref = f"{sheet.cell(row=1, column=status_column_index).coordinate}:{sheet.cell(row=len(target_df) + 1, column=status_column_index).coordinate}"

    # Speichere die aktualisierte Datei
    updated_file = target_file.replace(".xlsx", "_rowwise_updated.xlsx")
    book.save(updated_file)

    return updated_file


def main():
    inputs = gather_inputs()

    if inputs is None:
        print("Prozess abgebrochen.")
        return

    updated_file = compare_rows_and_update(
        inputs["target_file"],
        inputs["target_sheet"],
        inputs["source_file"],
        inputs["source_sheet"],
    )

    if updated_file:
        print("Aktualisierte Datei gespeichert unter:", updated_file)
    else:
        print("Der Vergleich wurde abgebrochen.")


# Hauptfunktion aufrufen
if __name__ == "__main__":
    main()
