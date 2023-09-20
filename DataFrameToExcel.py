import pandas as pd
import tkinter as tk
from openpyxl import Workbook
from tkinter import Listbox, MULTIPLE
from datetime import datetime

from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter.colorchooser import askcolor
from tkinter import (
    filedialog,
    messagebox,
    Tk,
    StringVar,
    OptionMenu,
    Label,
    Entry,
    Button,
    colorchooser,
    Toplevel,
    Radiobutton,
    E,
    W,
)
import os


# Usage of the class:
# Either call with params e.g

# from DataFrameToExcel import DataFrameToExcel as DFEClass
# app = DFEClaas(df, initial_file_name="default_name", show_gui=False)
#
# params = {
#    "sort_column": "project_number",
#    "title_bg_color": "000000",
#    "title_font_color": "FFFFFF",
#    "file_name": "output_test_file",
#    "correct_date_format": "no",
#    "highlight_rows": "no",
#    "highlight_column": "column_name",
#    "highlight_value": "column_content",
#    "highlight_color": "00FFFF",
#    "file_path": None,
#    "date_columns": ["date0", "date1",],
# }
#
## optional auto setting of file path
## directory_path = os.path.dirname(file_path)
## output_file_path = os.path.join(directory_path, params["file_name"] + ".xlsx")
## params["file_path"] = output_file_path
#
# app.save_to_excel(params=params)


# or call with GUI e.g
#
# from DataFrameToExcel import DataFrameToExcel as DFEClass
# app = DFEClass(df, master=root)


class DataFrameToExcel:
    def __init__(self, df, initial_file_name=None, master=None, show_gui=True):
        self.df = df
        self.initial_file_name = initial_file_name
        self.master = master
        
        if master is None:
            self.top = tk.Tk()
            print("I am master")
            self.top.title("DataFrame to Excel")
        else:
            self.top = tk.Toplevel(master)
            #self.top.withdraw()
            print("I am not master")
            self.top.title("DataFrame to Excel Settings")

        # Initialisieren Sie die Tkinter StringVar Attribute hier
        self.sort_column = StringVar(value=self.df.columns[0])
        self.title_bg_color = StringVar(value="000000")
        self.title_font_color = StringVar(value="FFFFFF")
        self.file_name = StringVar(
            value=self.initial_file_name if self.initial_file_name else "output"
        )
        self.correct_date_format = StringVar(value="no")
        self.highlight_rows = StringVar(value="no")
        self.highlight_column = StringVar(value=self.df.columns[0])
        self.highlight_value = StringVar(value="")
        self.highlight_color = StringVar(value="FFFFFF")

        if show_gui:
            self.setup_gui()

    def toggle_date_columns_listbox(self):
        if self.correct_date_format.get() == "yes":
            self.date_columns_listbox.grid(row=10, column=1, padx=20, pady=5)
        else:
            self.date_columns_listbox.grid_remove()

    def setup_gui(self):
        Label(self.top, text="Sort column:").grid(
            row=1, column=0, padx=20, pady=5, sticky=E
        )
        self.sort_column_menu = OptionMenu(
            self.top, self.sort_column, *self.df.columns
        )
        self.sort_column_menu.config(width=20)
        self.sort_column_menu.grid(row=1, column=1, padx=20, pady=5)

        Label(self.top, text="Title background color:").grid(
            row=2, column=0, padx=20, pady=5, sticky=E
        )
        Entry(self.top, textvariable=self.title_bg_color, width=20).grid(
            row=2, column=1, padx=20, pady=5, sticky=W
        )
        Button(
            self.top,
            text="Choose Color",
            command=lambda: self.choose_color(self.title_bg_color),
        ).grid(row=2, column=2, padx=20, pady=5)

        Label(self.top, text="Title font color:").grid(
            row=3, column=0, padx=20, pady=5, sticky=E
        )
        Entry(self.top, textvariable=self.title_font_color, width=20).grid(
            row=3, column=1, padx=20, pady=5, sticky=W
        )
        Button(
            self.top,
            text="Choose Color",
            command=lambda: self.choose_color(self.title_font_color),
        ).grid(row=3, column=2, padx=20, pady=5)

        Label(self.top, text="File name:").grid(
            row=4, column=0, padx=20, pady=5, sticky=E
        )
        Entry(self.top, textvariable=self.file_name, width=20).grid(
            row=4, column=1, padx=20, pady=5
        )

        Label(self.top, text="Correct date format:").grid(
            row=5, column=0, padx=20, pady=5, sticky=E
        )

        Radiobutton(
            self.top,
            text="Yes",
            variable=self.correct_date_format,
            value="yes",
            command=self.toggle_date_columns_listbox,
        ).grid(row=5, column=1, padx=20, pady=5, sticky=W)
        Radiobutton(
            self.top,
            text="No",
            variable=self.correct_date_format,
            value="no",
            command=self.toggle_date_columns_listbox,
        ).grid(row=5, column=1)

        Label(self.top, text="Highlight rows:").grid(
            row=6, column=0, padx=20, pady=5, sticky=E
        )
        Radiobutton(
            self.top, text="Yes", variable=self.highlight_rows, value="yes"
        ).grid(row=6, column=1, padx=20, pady=5, sticky=W)
        Radiobutton(
            self.top, text="No", variable=self.highlight_rows, value="no"
        ).grid(row=6, column=1)

        Label(self.top, text="Highlight column:").grid(
            row=7, column=0, padx=20, pady=5, sticky=E
        )
        self.highlight_column_menu = OptionMenu(
            self.top, self.highlight_column, *self.df.columns
        )
        self.highlight_column_menu.config(
            width=20
        )  # Setzen Sie die Breite des Dropdown-Menüs
        self.highlight_column_menu.grid(row=7, column=1, padx=20, pady=5)

        Label(self.top, text="Highlight value:").grid(
            row=8, column=0, padx=20, pady=5, sticky=E
        )
        Entry(self.top, textvariable=self.highlight_value, width=20).grid(
            row=8, column=1, padx=20, pady=5
        )

        Label(self.top, text="Highlight color:").grid(
            row=9, column=0, padx=20, pady=5, sticky=E
        )
        Entry(self.top, textvariable=self.highlight_color, width=20).grid(
            row=9, column=1, padx=20, pady=5, sticky=W
        )
        Button(
            self.top,
            text="Choose Color",
            command=lambda: self.choose_color(self.highlight_color),
        ).grid(row=9, column=2, padx=20, pady=5)

        self.date_columns_listbox = Listbox(
            self.top, selectmode=MULTIPLE, exportselection=0, width=20, height=4
        )
        for col in self.df.columns:
            self.date_columns_listbox.insert("end", col)

        if self.correct_date_format.get() == "yes":
            Label(self.top, text="Date columns:").grid(
                row=10, column=0, padx=20, pady=5, sticky=E
            )
            self.date_columns_listbox.grid(row=10, column=1, padx=20, pady=5)

        # Fügen Sie eine Schaltfläche hinzu, um die save_to_excel Methode aufzurufen
        save_button = Button(
            self.top, text="Save to Excel", command=self.save_to_excel
        )
        save_button.grid(row=13, column=0, columnspan=2, pady=20)

        # Starten Sie die Tkinter-Hauptschleife
        #self.top.mainloop()

    def choose_color(self, var):
        color_code = askcolor()[1]
        if color_code:
            var.set(color_code.lstrip("#"))

    def get_selected_date_columns(self):
        return [self.df.columns[i] for i in self.date_columns_listbox.curselection()]

    def get_save_path(self):

        """Rückgabe des Pfads, in dem die Datei gespeichert wurde."""
        try:
            if self.select_xlsx_path:
                return self.select_xlsx_path
            else:
                raise ValueError("Der Pfad ist nicht gesetzt.")
        except ValueError as e:
            print(f"Fehler: {e}")
            self.select_xlsx_path = None

    def save_to_excel(self, params=None):
        if params:
            # print(
            #    f"Parameters received: {params}"
            # )  # Drucken Sie die übergebenen Parameter

            self.sort_column.set(params.get("sort_column", self.df.columns[0]))
            self.title_bg_color.set(params.get("title_bg_color", "000000"))
            self.title_font_color.set(params.get("title_font_color", "FFFFFF"))
            self.file_name.set(params.get("file_name", "output"))
            self.correct_date_format.set(params.get("correct_date_format", "yes"))
            self.highlight_rows.set(params.get("highlight_rows", "no"))
            self.highlight_column.set(
                params.get("highlight_column", self.df.columns[0])
            )
            self.highlight_value.set(params.get("highlight_value", ""))
            self.highlight_color.set(params.get("highlight_color", "FFFFFF"))
            date_columns = params.get("date_columns", [])
            # if isinstance(date_columns, str):
            #    date_columns = [date_columns]

            # optional print test
            # print(
            #    f"Parameter correct date format gesetzt: {self.correct_date_format.get()}"
            # )
            # print(f"Date columns: {date_columns}")

            save_path = params.get("file_path")
            if not save_path:
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=self.file_name.get(),
                    filetypes=[("Excel files", "*.xlsx")],
                )
                if not save_path:
                    messagebox.showerror("Error", "No save location selected.")
                    return

        else:
            # Get save path through GUI
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.file_name.get(),
                filetypes=[("Excel files", "*.xlsx")],
            )
            if not save_path:
                messagebox.showerror("Error", "No save location selected.")
                return

        # Sort DataFrame
        self.df.sort_values(self.sort_column.get(), inplace=True)    
        
        try:
            # Write to Excel
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                self.df.to_excel(writer, sheet_name="Sheet", index=False)
                workbook = writer.book
                worksheet = workbook["Sheet"]
                self.select_xlsx_path = save_path

                for cell in worksheet[1:1]:
                    cell.font = Font(color=self.title_font_color.get())
                    cell.fill = PatternFill(
                        fgColor=self.title_bg_color.get(), fill_type="solid"
                    )
                    cell.border = Border(
                        left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"),
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                if params is None:
                    date_columns = self.get_selected_date_columns()
                else:
                    date_columns = params.get("date_columns")

                # print(f"Date columns: {date_columns}")
                if self.correct_date_format.get() == "yes":
                    date_format = "DD.MM.YYYY"
                    for date_col in date_columns:
                        col_idx = self.df.columns.get_loc(date_col)
                        col_letter = chr(ord("A") + col_idx)
                        for cell in worksheet[col_letter]:
                            if cell.row == 1:
                                continue
                            # if isinstance(cell.value, pd.Timestamp):
                            cell.number_format = date_format

                if self.highlight_rows.get() == "yes":
                    highlight_fill = PatternFill(
                        start_color=self.highlight_color.get(),
                        end_color=self.highlight_color.get(),
                        fill_type="solid",
                    )
                    highlight_col_idx = self.df.columns.get_loc(
                        self.highlight_column.get()
                    )

                    for row_idx, row in enumerate(
                        dataframe_to_rows(self.df, index=False, header=False), 2
                    ):
                        if str(row[highlight_col_idx]) == self.highlight_value.get():
                            for cell in worksheet[row_idx]:
                                cell.fill = highlight_fill

                # Apply formatting
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[
                        column[0].column_letter
                    ].width = adjusted_width

            messagebox.showinfo("Success", "File saved successfully.")
            
        except Exception as e:
            # import traceback
            #
            # traceback.print_exc()  # Diese Zeile wird den vollständigen Stack-Trace ausgeben
            messagebox.showerror("Error", f"Failed to save file: {e}")
        #global select_xlsx_path
        select_xlsx_path = self.get_save_path()
        print(f"Datei gespeichert: {select_xlsx_path}")
        self.top.destroy()
        return select_xlsx_path

    def close(self):
        self.top.destroy() 
